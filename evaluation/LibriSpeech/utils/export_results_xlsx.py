#!/usr/bin/env python3
"""paper_result/ASR의 mode × split 실행 결과를 엑셀 한 파일로 모은다.

사용 예:
    python evaluation/LibriSpeech/utils/export_results_xlsx.py --tag c16_run01

시트 구성
    요약     — 실행 12개를 한 줄씩. WER/S/I/D, FSL 분위수, FTL, 커밋 사유 분포
    피벗     — mode × split 행렬 (WER / FSL avg / FTL / finish 커밋)
    실행설정 — meta.json에서 뽑은 실행 조건. 나중에 "이 표가 어떤 조건이었나"를 여기서 본다

S/I/D는 러너와 같은 정규화(compute_wer_for_rows의 normalize)를 써서 metric.json의 WER과
어긋나지 않게 한다. 다른 정규화를 쓰면 0.1%p 수준의 설명 안 되는 차이가 생긴다.
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import jiwer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SPLITS = [("testclean", "test-clean"), ("testother", "test-other"),
          ("devclean", "dev-clean"), ("devother", "dev-other")]
MODES = [(2, "always-commit", "baseline(1.0.0)"),
         (3, "dot-commit", "baseline(1.0.0)"),
         (4, "seg-commit", "en-silence-c80")]
REASONS = ["always", "dot", "seg", "vad", "finish"]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def normalize(text: str) -> str:
    """test_qwen3_librispeech.compute_wer_for_rows와 동일한 정규화."""
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text)
    return " ".join(text.split())


def percentile(values, p):
    values = sorted(values)
    if not values:
        return None
    k = (len(values) - 1) * p / 100
    f = int(k)
    c = min(f + 1, len(values) - 1)
    return values[f] + (values[c] - values[f]) * (k - f)


def sid_rates(rows):
    """치환/삽입/삭제를 참조 단어 수로 나눈 비율. jiwer로 한 번에 처리한다."""
    pairs = [(normalize(r["reference"]), normalize(r["hypothesis"]))
             for r in rows if r.get("reference") and r.get("hypothesis")]
    pairs = [(r, h) for r, h in pairs if r.strip() and h.strip()]
    if not pairs:
        return None, None, None
    refs, hyps = zip(*pairs)
    out = jiwer.process_words(list(refs), list(hyps))
    n = sum(len(r.split()) for r in refs)
    return out.substitutions / n, out.insertions / n, out.deletions / n


def collect(root: Path, tag: str):
    runs = []
    for mode, policy, model in MODES:
        for key, label in SPLITS:
            run_dir = root / f"mode{mode}" / "full" / f"{key}_{tag}"
            metric = run_dir / "metric.json"
            if not metric.exists():
                print(f"  건너뜀 (metric.json 없음): {run_dir}")
                continue
            d = json.loads(metric.read_text())
            o, conc = d["overall"], d.get("concurrency", {})
            cs = o["commit_stats"]
            raw = d["raw_results"]

            fsl = [x["fsl_sec"] for r in raw for x in (r.get("segment_metrics") or [])
                   if x.get("fsl_sec") is not None]
            s, i, dele = sid_rates(raw)
            meta_path = run_dir / "meta.json"
            meta = json.loads(meta_path.read_text()) if meta_path.exists() else {}

            runs.append({
                "mode": f"mode{mode}", "policy": policy, "model": model, "split": label,
                "files": o["num_files"], "wer": o["wer"], "s": s, "i": i, "d": dele,
                "fsl_avg": o["avg_fsl_sec"],
                "fsl_p50": percentile(fsl, 50), "fsl_p95": percentile(fsl, 95),
                "fsl_p99": percentile(fsl, 99),
                "ftl": o["first_token_latency"], "runtime": o["model_runtime"],
                "commits": cs["total"], "tok_per_commit": o["avg_output_tokens_per_commit"],
                **{f"c_{r}": cs["counts"].get(r, 0) for r in REASONS},
                "makespan": conc.get("makespan_sec"),
                "efficiency": (conc.get("queue_balance_prediction") or {}).get("predicted_efficiency"),
                "meta": meta,
            })
    return runs


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def autofit(ws, min_w=9, max_w=44):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, min_w), max_w)


def sheet_summary(wb, runs):
    ws = wb.create_sheet("요약")
    cols = [
        ("모드", "mode", None), ("정책", "policy", None), ("모델", "model", None),
        ("split", "split", None), ("파일", "files", "#,##0"),
        ("WER", "wer", "0.00%"), ("치환 S", "s", "0.00%"), ("삽입 I", "i", "0.00%"),
        ("삭제 D", "d", "0.00%"),
        ("FSL avg", "fsl_avg", "0.0000"), ("FSL p50", "fsl_p50", "0.0000"),
        ("FSL p95", "fsl_p95", "0.0000"), ("FSL p99", "fsl_p99", "0.0000"),
        ("FTL", "ftl", "0.000"), ("model_runtime", "runtime", "0.000"),
        ("커밋 수", "commits", "#,##0"), ("tok/커밋", "tok_per_commit", "0.00"),
        ("always", "c_always", "#,##0"), ("dot", "c_dot", "#,##0"), ("seg", "c_seg", "#,##0"),
        ("vad", "c_vad", "#,##0"), ("finish", "c_finish", "#,##0"),
        ("makespan(s)", "makespan", "0"), ("큐 효율", "efficiency", "0.0%"),
    ]
    ws.append([c[0] for c in cols])
    for r in runs:
        ws.append([r.get(key) for _, key, _ in cols])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cols)):
        for idx, cell in enumerate(row):
            cell.border = BORDER
            fmt = cols[idx][2]
            if fmt:
                cell.number_format = fmt
    style_header(ws, len(cols))
    autofit(ws)
    return ws


def sheet_pivot(wb, runs):
    ws = wb.create_sheet("피벗")
    lookup = {(r["mode"], r["split"]): r for r in runs}
    labels = [lbl for _, lbl in SPLITS]
    blocks = [("WER", "wer", "0.00%"), ("FSL avg (s)", "fsl_avg", "0.0000"),
              ("FSL p99 (s)", "fsl_p99", "0.0000"), ("FTL (s)", "ftl", "0.000"),
              ("finish 커밋", "c_finish", "#,##0")]
    row = 1
    for title, key, fmt in blocks:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1
        header = row
        ws.cell(row=row, column=1, value="모드")
        for c, lbl in enumerate(labels, start=2):
            ws.cell(row=row, column=c, value=lbl)
        for c in range(1, len(labels) + 2):
            cell = ws.cell(row=header, column=c)
            cell.fill, cell.font = HEAD_FILL, HEAD_FONT
            cell.alignment = Alignment(horizontal="center")
        row += 1
        for mode, _, _ in MODES:
            ws.cell(row=row, column=1, value=f"mode{mode}").font = Font(bold=True)
            for c, lbl in enumerate(labels, start=2):
                r = lookup.get((f"mode{mode}", lbl))
                cell = ws.cell(row=row, column=c, value=r.get(key) if r else None)
                cell.number_format = fmt
                cell.border = BORDER
            row += 1
        row += 1
    autofit(ws, min_w=13)
    return ws


def sheet_config(wb, runs):
    ws = wb.create_sheet("실행설정")
    ws.append(["모드", "split", "서버 model", "always_commit", "enable_dot_commit",
               "rep_dedup", "no_vad", "chunk(ms)", "send(ms)", "trailing(ms)",
               "클라이언트", "번역", "실행 시각"])
    for r in runs:
        meta = r["meta"]
        cli = meta.get("cli_args", {})
        sc = meta.get("server_config", {})
        ws.append([
            r["mode"], r["split"], sc.get("model"),
            sc.get("always_commit"), sc.get("enable_dot_commit"),
            sc.get("rep_dedup"), sc.get("no_vad"),
            cli.get("chunk_size_ms"), cli.get("send_interval_ms"),
            cli.get("trailing_silence_ms"), cli.get("num_clients"),
            "off" if not cli.get("target_lang") else cli.get("target_lang"),
            (meta.get("timestamp") or "")[:19].replace("T", " "),
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=13):
        for cell in row:
            cell.border = BORDER
    style_header(ws, 13)
    autofit(ws)
    return ws


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--root", default="evaluation/LibriSpeech/paper_result/ASR")
    ap.add_argument("--tag", default="c16_run01")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    root = Path(args.root)
    out = Path(args.out) if args.out else root / f"mode234_full_{args.tag}.xlsx"

    print(f"수집 중: {root} (tag={args.tag})")
    runs = collect(root, args.tag)
    if not runs:
        raise SystemExit("실행 결과를 찾지 못했습니다.")

    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, runs)
    sheet_pivot(wb, runs)
    sheet_config(wb, runs)
    wb.save(out)
    print(f"저장: {out}  (실행 {len(runs)}건)")


if __name__ == "__main__":
    main()
